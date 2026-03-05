#!/usr/bin/env python3
"""
TermIndex Knowledge Base — Document Ingestion Pipeline

Accepts a PDF or XLSX file, extracts text, chunks it into ~500-token segments
with configurable overlap, and outputs JSON chunks with metadata.

Usage:
    python ingest.py --input report.pdf --output chunks/ --standard pcaf
    python ingest.py --input data.xlsx --output chunks/ --standard pcaf
"""

import argparse
import json
import os
import re
import sys
import uuid
from pathlib import Path
from typing import Optional


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from a PDF file using PyPDF2 or pdfplumber."""
    text = ""

    # Try pdfplumber first (better table extraction)
    try:
        import pdfplumber

        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n\n"
        if text.strip():
            return text
    except ImportError:
        pass

    # Fallback to PyPDF2
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(file_path)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n\n"
        return text
    except ImportError:
        pass

    print(
        "ERROR: Neither pdfplumber nor PyPDF2 is installed. "
        "Install one with: pip install pdfplumber  OR  pip install PyPDF2",
        file=sys.stderr,
    )
    sys.exit(1)


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from an XLSX file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "ERROR: openpyxl is not installed. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    text_parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===\n")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text_parts.append(row_text)

    wb.close()
    return "\n".join(text_parts)


def estimate_tokens(text: str) -> int:
    """Estimate token count (~4 characters per token for English text)."""
    return len(text) // 4


def chunk_text(
    text: str,
    chunk_size: int = 500,
    chunk_overlap: int = 50,
) -> list[dict]:
    """
    Split text into chunks of approximately chunk_size tokens
    with chunk_overlap token overlap between consecutive chunks.

    Returns a list of dicts with 'text', 'start_char', 'end_char'.
    """
    # Approximate characters per token
    chars_per_token = 4
    target_chars = chunk_size * chars_per_token
    overlap_chars = chunk_overlap * chars_per_token

    # Split into paragraphs first, then reassemble into chunks
    paragraphs = re.split(r"\n\s*\n", text)
    paragraphs = [p.strip() for p in paragraphs if p.strip()]

    chunks = []
    current_chunk = ""
    current_start = 0
    char_position = 0

    for para in paragraphs:
        para_with_sep = para + "\n\n"

        if len(current_chunk) + len(para_with_sep) > target_chars and current_chunk:
            # Save current chunk
            chunks.append(
                {
                    "text": current_chunk.strip(),
                    "start_char": current_start,
                    "end_char": current_start + len(current_chunk.strip()),
                }
            )

            # Start new chunk with overlap
            if overlap_chars > 0 and len(current_chunk) > overlap_chars:
                overlap_text = current_chunk[-overlap_chars:]
                current_chunk = overlap_text + para_with_sep
            else:
                current_chunk = para_with_sep
            current_start = char_position
        else:
            if not current_chunk:
                current_start = char_position
            current_chunk += para_with_sep

        char_position += len(para_with_sep)

    # Don't forget the last chunk
    if current_chunk.strip():
        chunks.append(
            {
                "text": current_chunk.strip(),
                "start_char": current_start,
                "end_char": current_start + len(current_chunk.strip()),
            }
        )

    return chunks


def create_chunk_json(
    chunk_data: dict,
    index: int,
    standard_id: str,
    source_file: str,
    total_chunks: int,
) -> dict:
    """Create a structured JSON chunk with metadata."""
    chunk_id = f"{standard_id}_ingest_{index + 1:04d}"

    return {
        "chunk_id": chunk_id,
        "standard_id": standard_id,
        "source_file": os.path.basename(source_file),
        "section": "ingested",
        "topic": f"Ingested chunk {index + 1} of {total_chunks}",
        "language": "en",
        "content": chunk_data["text"],
        "metadata": {
            "chunk_index": index,
            "total_chunks": total_chunks,
            "start_char": chunk_data["start_char"],
            "end_char": chunk_data["end_char"],
            "estimated_tokens": estimate_tokens(chunk_data["text"]),
            "keywords": [],
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="TermIndex Knowledge Base — Document Ingestion Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python ingest.py --input report.pdf --output chunks/ --standard pcaf
    python ingest.py --input data.xlsx --output chunks/ --standard pcaf --chunk-size 300
        """,
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to input file (PDF or XLSX)",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output directory for JSON chunks",
    )
    parser.add_argument(
        "--standard",
        default="pcaf",
        help="Standard ID (default: pcaf)",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=500,
        help="Target chunk size in tokens (default: 500)",
    )
    parser.add_argument(
        "--chunk-overlap",
        type=int,
        default=50,
        help="Overlap between chunks in tokens (default: 50)",
    )

    args = parser.parse_args()

    # Validate input file
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    suffix = input_path.suffix.lower()
    if suffix not in (".pdf", ".xlsx"):
        print(
            f"ERROR: Unsupported file type '{suffix}'. Supported: .pdf, .xlsx",
            file=sys.stderr,
        )
        sys.exit(1)

    # Create output directory
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Extract text
    print(f"Extracting text from {input_path.name}...")
    if suffix == ".pdf":
        text = extract_text_from_pdf(str(input_path))
    else:
        text = extract_text_from_xlsx(str(input_path))

    if not text.strip():
        print("WARNING: No text extracted from input file.", file=sys.stderr)
        sys.exit(1)

    total_tokens = estimate_tokens(text)
    print(f"Extracted ~{total_tokens} tokens from {input_path.name}")

    # Chunk text
    print(
        f"Chunking with size={args.chunk_size}, overlap={args.chunk_overlap}..."
    )
    raw_chunks = chunk_text(text, args.chunk_size, args.chunk_overlap)
    print(f"Generated {len(raw_chunks)} chunks")

    # Write chunks as JSON
    chunk_index = []
    for i, raw_chunk in enumerate(raw_chunks):
        chunk_json = create_chunk_json(
            raw_chunk,
            i,
            args.standard,
            str(input_path),
            len(raw_chunks),
        )

        # Write individual chunk file
        chunk_file = output_dir / f"{chunk_json['chunk_id']}.json"
        with open(chunk_file, "w", encoding="utf-8") as f:
            json.dump(chunk_json, f, indent=2, ensure_ascii=False)

        # Add to index (without content)
        index_entry = {k: v for k, v in chunk_json.items() if k != "content"}
        chunk_index.append(index_entry)

    # Write chunk index
    index_file = output_dir / "chunk_index.json"
    with open(index_file, "w", encoding="utf-8") as f:
        json.dump(chunk_index, f, indent=2, ensure_ascii=False)

    print(f"Written {len(raw_chunks)} chunks to {output_dir}/")
    print(f"Chunk index written to {index_file}")


if __name__ == "__main__":
    main()
